# Import all libraries
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import pandas as pd
import pyttsx3


# ------------------- Step 1: Setup Face Encoding -------------------

# Folder containing known faces
path = 'images'

images = []
classnames = []

# Get all files from the images folder
mylist = os.listdir(path)

for cl in mylist:
    curImg = cv2.imread(os.path.join(path, cl))

    if curImg is not None:
        images.append(curImg)
        classnames.append(os.path.splitext(cl)[0])


# Function to encode known images
def findEncodings(images):
    encodeList = []

    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

        encodings = face_recognition.face_encodings(img)

        if encodings:
            encodeList.append(encodings[0])

    return encodeList


encodeListKnown = findEncodings(images)

print("[INFO] Encoding Complete")


# ------------------- Step 2: Setup Excel Workbook -------------------

# Attendance marking function
def markAttendance(name):
    file_name = 'Attendance.xlsx'

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Date", "Time"])

    now = datetime.now()

    date = now.strftime('%Y-%m-%d')
    time = now.strftime('%H:%M:%S')

    # Get names already present in the sheet
    names_in_sheet = [
        cell.value for cell in sheet['A']
        if cell.value != "Name"
    ]

    # Check whether attendance is already marked
    if name not in names_in_sheet:
        sheet.append([name, date, time])
        workbook.save(file_name)

        print(f"[INFO] Attendance marked for {name}")

    else:
        print(f"[INFO] {name} already marked.")


# ------------------- Step 3: Setup Text-to-Speech -------------------

engine = pyttsx3.init()


def speak(text):
    engine.say(text)
    engine.runAndWait()


# ------------------- Step 4: Smart Attendance Logic -------------------

def smart_attendance():

    cap = cv2.VideoCapture(0)

    already_marked = set()

    while True:

        success, img = cap.read()

        if not success:
            print("[ERROR] Unable to access webcam.")
            break

        # Resize image for faster face recognition
        imgS = cv2.resize(
            img,
            (0, 0),
            None,
            0.25,
            0.25
        )

        imgS = cv2.cvtColor(
            imgS,
            cv2.COLOR_BGR2RGB
        )

        # Detect faces
        facesCurFrame = face_recognition.face_locations(imgS)

        # Encode detected faces
        encodesCurFrame = face_recognition.face_encodings(
            imgS,
            facesCurFrame
        )

        # Compare each detected face
        for encodeFace, faceLoc in zip(
            encodesCurFrame,
            facesCurFrame
        ):

            matches = face_recognition.compare_faces(
                encodeListKnown,
                encodeFace
            )

            faceDis = face_recognition.face_distance(
                encodeListKnown,
                encodeFace
            )

            # Make sure known faces are available
            if len(faceDis) == 0:
                continue

            matchIndex = np.argmin(faceDis)

            # Convert face coordinates back to original image size
            y1, x2, y2, x1 = faceLoc

            y1 *= 4
            x2 *= 4
            y2 *= 4
            x1 *= 4

            # Check whether the face is recognized
            if (
                matches[matchIndex]
                and faceDis[matchIndex] < 0.6
            ):

                name = classnames[matchIndex].upper()

                # Draw green rectangle
                cv2.rectangle(
                    img,
                    (x1, y1),
                    (x2, y2),
                    (0, 255, 0),
                    2
                )

                # Display name
                cv2.putText(
                    img,
                    name,
                    (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,
                    (255, 255, 255),
                    2
                )

                # Mark attendance only once per session
                if name not in already_marked:

                    markAttendance(name)

                    already_marked.add(name)

                    speak(
                        f"Attendance marked for {name}"
                    )

            else:

                # Draw red rectangle for unknown person
                cv2.rectangle(
                    img,
                    (x1, y1),
                    (x2, y2),
                    (0, 0, 255),
                    2
                )

                cv2.putText(
                    img,
                    "New Student Detected",
                    (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    1,
                    (0, 0, 255),
                    2
                )

        # Display webcam
        cv2.imshow(
            'Webcam - Press Q to Exit',
            img
        )

        # Press Q to exit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


# ------------------- Step 5: Setup Tkinter GUI -------------------

def main_gui():

    window = tk.Tk()

    window.title("Smart Attendance System")

    window.configure(bg="#f0f0f0")

    window.geometry("400x300")

    window.resizable(False, False)

    # Center window
    window.update_idletasks()

    width = window.winfo_width()
    height = window.winfo_height()

    x = (
        window.winfo_screenwidth() // 2
    ) - (
        width // 2
    )

    y = (
        window.winfo_screenheight() // 2
    ) - (
        height // 2
    )

    window.geometry(
        f"{width}x{height}+{x}+{y}"
    )

    # Title label
    title_label = tk.Label(
        window,
        text="Smart Attendance System",
        font=("Helvetica", 16, "bold"),
        bg="#f0f0f0",
        fg="#333"
    )

    title_label.pack(pady=20)

    # Style buttons
    style = ttk.Style()

    style.configure(
        "TButton",
        font=("Helvetica", 12),
        padding=10
    )

    # View attendance function
    def open_attendance():

        try:

            df = pd.read_excel(
                'Attendance.xlsx'
            )

            messagebox.showinfo(
                "Attendance Sheet",
                df.to_string(index=False)
            )

        except FileNotFoundError:

            messagebox.showerror(
                "Error",
                "Attendance file not found."
            )

    # Start attendance button
    ttk.Button(
        window,
        text="Start Smart Attendance",
        command=smart_attendance
    ).pack(pady=10)

    # View attendance button
    ttk.Button(
        window,
        text="View Attendance",
        command=open_attendance
    ).pack(pady=10)

    # Exit button
    ttk.Button(
        window,
        text="Exit",
        command=window.destroy
    ).pack(pady=10)

    window.mainloop()


# ------------------- Step 6: Run Main GUI -------------------

if _name_ == "_main_":
    main_gui()